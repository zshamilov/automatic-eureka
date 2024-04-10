from datetime import datetime, timedelta

import pandas as pd

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List

from products.Decision.framework.model import DiagramInOutParameterFullViewDto, ParameterType
from products.Decision.utilities.custom_models import IntValueType, BasicPrimitiveValues


def tst_report_validator_values(report_path, page, test_set_name=None, test_set_count=None,
                                test_set_successes=None, test_set_failures=None, test_set_result=None,
                                out_var_expectation=None, out_var_fact=None,
                                diagram_exec_status_expectation=None, diagram_exec_status_fact=None):
    file_valid = False
    if page == 1:
        report = pd.read_excel(report_path)
        ts_name = str(report.iat[0, 1])
        ts_count = int(report.iat[1, 1])
        ts_successes = int(report.iat[2, 1])
        ts_failures = int(report.iat[3, 1])
        ts_result = str(report.iat[7, 1])
        if ts_name == test_set_name and ts_count == test_set_count \
                and ts_successes == test_set_successes and ts_failures == test_set_failures \
                and ts_result == test_set_result:
            file_valid = True
    if page == 2:
        report = pd.read_excel(report_path, sheet_name="Тестовый набор 1")
        ov_expectation = int(report.iat[0, 1])
        ov_fact = int(report.iat[0, 2])
        des_expectation = str(report.iat[1, 1])
        des_fact = str(report.iat[1, 2])
        if ov_expectation == out_var_expectation and ov_fact == out_var_fact \
                and des_expectation == diagram_exec_status_expectation \
                and des_fact == diagram_exec_status_fact:
            file_valid = True
    return file_valid


def tst_report_validator_titles(report_path, page, test_set_name=None, test_set_count=None,
                                test_set_successes=None, test_set_failures=None, test_set_result=None,
                                out_var_title=None, diagram_exec_status_title=None):
    file_valid = False
    if page == 1:
        report = pd.read_excel(report_path)
        ts_name = str(report.iat[0, 0])
        ts_count = str(report.iat[1, 0])
        ts_successes = str(report.iat[2, 0])
        ts_failures = str(report.iat[3, 0])
        ts_result = str(report.iat[7, 0])
        if ts_name == test_set_name and ts_count == test_set_count \
                and ts_successes == test_set_successes and ts_failures == test_set_failures \
                and ts_result == test_set_result:
            file_valid = True
    if page == 2:
        report = pd.read_excel(report_path, sheet_name="Тестовый набор 1")
        ov_title = str(report.iat[0, 0])
        des_title = str(report.iat[1, 0])
        if ov_title == out_var_title and des_title == diagram_exec_status_title:
            file_valid = True
    return file_valid


def file_testing_get_inout_variables(file_path):
    out_variables_header_name = 'Ожидаемые результаты'
    in_variables = []
    out_variables = []
    out_variables_header_cell = None
    wb = load_workbook(file_path)
    sheet: Worksheet = wb['Тестовые наборы']
    complex_type_cells = []
    ctype_with_values = dict()

    # Поиск ячейки с заголовков выходной переменной
    for row in sheet.iter_rows():
        for cell in row:
            cell: Cell
            if cell.value == out_variables_header_name:
                out_variables_header_cell = cell

    current_cell: Cell = sheet['B2']
    # Собираем переменные до пустой ячейки
    while current_cell.value is not None:
        # если ячейка содержит ссылку => это комплексная переменная
        if current_cell.hyperlink is not None:
            complex_type_cells.append(current_cell)
        # если ячейка находится до выходных переменных => она входная
        if current_cell.col_idx < out_variables_header_cell.col_idx:
            in_variables.append(current_cell.value.removeprefix("_").removesuffix("[]"))
        else:
            out_variables.append(current_cell.value.removeprefix("_res_").removesuffix("[]"))
        current_cell: Cell = sheet.cell(row=current_cell.row,
                                        column=current_cell.column + 1)

    for cell in complex_type_cells:
        ctype_name = str(cell.value).removesuffix("[]")
        ctype_sheet_name = ctype_name.removeprefix("_")
        ctype_with_values[ctype_name] = []
        ctype_sheet: Worksheet = wb[ctype_sheet_name]
        current_ctype_cell: Cell = ctype_sheet['B2']
        while current_ctype_cell.value is not None:
            ctype_with_values[ctype_name].append(current_ctype_cell.value)
            current_ctype_cell = ctype_sheet.cell(row=current_ctype_cell.row,
                                                  column=current_ctype_cell.column + 1)

    return {"in_variables": in_variables,
            "out_variables": out_variables,
            "ctype_with_values": ctype_with_values}


def file_testing_generate_test_cases(file_path,
                                     diagram_variables: List[DiagramInOutParameterFullViewDto],
                                     test_to_fail=False, custom_values: dict = None) -> list[tuple]:
    """
    Функция для генерации тестов на основе переменных диаграммы.

    :param file_path: Путь к заполняемому файлу.
    :param diagram_variables: Список переменных диаграммы List[DiagramInOutParameterFullViewDto],
        которые будут использоваться в тестах.
    :param test_to_fail: Параметр, определяющий, должен ли тест провалиться. Если True, изменяет ВСЕ ожидаемые
                         значения так, что тест не пройдет. По умолчанию False.
    :param custom_values: Словарь с пользовательскими значениями для переменных формата {имя_перемнной:Кастомное значение}.
        ВАЖНО! для типов дата, дата_время, время значения должны быть в типах библиотеки datetime

    Функция дублирует сквозные переменные диаграммы(если они есть), инициализирует базовые и ожидаемые значения для каждой из
    переменных на основе их типов. Затем функция открывает указанный файл и обрабатывает его, заполняя ячейки
    заданными в параметре custom_values или стандартными значениями переменных.
    """
    if custom_values is None:
        custom_values = dict()

    # убираем ненужный для теста тип из класса
    primitive_types_without_rule = [type_id for type_id in IntValueType
                                    if type_id != IntValueType.complex_type_rule]

    basic_values = {primitive_type.value: basic_value.value
                    for primitive_type, basic_value
                    in zip(primitive_types_without_rule, BasicPrimitiveValues)}

    basic_values[IntValueType.date] = datetime.strptime(basic_values[IntValueType.date], '%Y-%m-%d')
    basic_values[IntValueType.dateTime] = datetime.strptime(basic_values[IntValueType.dateTime], '%Y-%m-%d %H:%M:%S.%f')
    basic_values[IntValueType.time] = datetime.strptime(basic_values[IntValueType.time], '%H:%M:%S').time()

    expected_values = basic_values.copy()

    # подменяем все значения если тест-кейс должен упасть
    if test_to_fail:
        for primitive_type_id in expected_values:
            if primitive_type_id in (IntValueType.int.value, IntValueType.float.value,
                                     IntValueType.long.value, IntValueType.str.value):
                expected_values[primitive_type_id] = expected_values[primitive_type_id] * 2
            elif primitive_type_id in (IntValueType.date.value, IntValueType.dateTime.value):
                expected_values[primitive_type_id] = expected_values[primitive_type_id] + timedelta(days=1)

            # требует исправления, но бессмысленно до фикса DEV-21601
            # elif primitive_type_id == IntValueType.time.value:
            #     expected_values[primitive_type_id] = expected_values[primitive_type_id] + timedelta(hours=1)

            else:
                expected_values[primitive_type_id] = not (expected_values[primitive_type_id])

    out_variables_header_name = 'Ожидаемые результаты'
    out_variables_header_cell = None
    wb: Workbook = load_workbook(file_path)
    sheet: Worksheet = wb['Тестовые наборы']

    # Поиск ячейки с заголовком выходной переменной
    for row in sheet.iter_rows():
        for cell in row:
            cell: Cell
            if cell.value == out_variables_header_name:
                out_variables_header_cell = cell

    # ставим номер тест кейса
    sheet['A3'].value = 1

    values_of_tests = []
    current_cell: Cell = sheet['B2']

    current_d_var = None

    # заполняем значения в столбцах с именем переменной кастомным значением, если передано или базовым.
    while current_cell.value is not None:
        current_d_var = [d_var for d_var
                         in diagram_variables
                         if d_var.parameterName == current_cell.value].pop()

        if current_cell.column < out_variables_header_cell.column:
            sheet.cell(row=current_cell.row + 1, column=current_cell.column).value = (
                    custom_values[current_d_var.parameterName]
                    if current_d_var.parameterName in custom_values
                    else basic_values[current_d_var.typeId])
        else:

            if current_d_var.parameterName == 'diagram_execute_status':
                sheet.cell(row=current_cell.row + 1, column=current_cell.column).value = (
                    custom_values[current_d_var.parameterName]
                    if current_d_var.parameterName in custom_values
                    else 1)
            else:
                sheet.cell(row=current_cell.row + 1, column=current_cell.column).value = (
                    custom_values[current_d_var.parameterName]
                    if current_d_var.parameterName in custom_values
                    else expected_values[current_d_var.typeId])

        values_of_tests.append((current_cell.value,
                                sheet.cell(row=current_cell.row + 1, column=current_cell.column).value))
        current_cell = sheet.cell(row=current_cell.row, column=current_cell.column + 1)

    wb.save(file_path)

    #
    # for cell in complex_type_cells:
    #     ctype_sheet_name = str(cell.value).removeprefix("_").removesuffix("[]")
    #     ctype_with_values[ctype_sheet_name] = []
    #     ctype_sheet: Worksheet = wb[ctype_sheet_name]
    #     current_ctype_cell: Cell = ctype_sheet['B2']
    #     while current_ctype_cell.value is not None:
    #         ctype_with_values[ctype_sheet_name].append(current_ctype_cell.value)
    #         current_ctype_cell = ctype_sheet.cell(row=current_ctype_cell.row,
    #                                               column=current_ctype_cell.column + 1)

    return values_of_tests
